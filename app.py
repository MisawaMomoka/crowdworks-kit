"""
app.py - クラウドワークス案件獲得自動化キット
Streamlit アプリ メインファイル

ローカル起動: streamlit run app.py
"""

import os

import streamlit as st
from scraper import run_scraping
from scorer import score_all_jobs
from proposal_generator import generate_all_proposals, generate_proposal
import pandas as pd
import io
import threading
from streamlit.runtime.scriptrunner import add_script_run_ctx, get_script_run_ctx

# ======================================================================
# ページ設定
# ======================================================================

st.set_page_config(
    page_title="CW案件獲得自動化キット",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    /* フォントと余白の調整 */
    .block-container { padding-top: 2rem; }
    h1 { font-size: 1.6rem; font-weight: 700; }
    h2 { font-size: 1.2rem; font-weight: 600; }
    h3 { font-size: 1.0rem; font-weight: 600; }

    /* スコアバッジ */
    .score-high { color: #1a7f4b; font-weight: 700; }
    .score-mid  { color: #b07d00; font-weight: 700; }
    .score-low  { color: #c0392b; font-weight: 700; }

    /* テーブル */
    .stDataFrame { font-size: 0.85rem; }

    /* ボタン */
    .stButton > button { font-size: 0.9rem; }
</style>
""", unsafe_allow_html=True)

# ======================================================================
# セッション初期化
# ======================================================================

if "scored_results" not in st.session_state:
    st.session_state.scored_results = []
if "proposals_done" not in st.session_state:
    st.session_state.proposals_done = False

# ======================================================================
# サイドバー
# ======================================================================

with st.sidebar:
    st.title("設定")

    st.subheader("クラウドワークス ログイン情報")
    cw_email = st.text_input("メールアドレス", placeholder="example@gmail.com")
    cw_password = st.text_input("パスワード", type="password")

    st.subheader("検索キーワード")
    keywords_str = st.text_area(
        "1行に1キーワード",
        value="データ入力\nメール返信代行\n継続案件\n事務",
        height=110,
    )
    keywords = [k.strip() for k in keywords_str.splitlines() if k.strip()]

    max_pages = st.slider("各キーワードの取得ページ数", 1, 10, 2)
    threshold = st.slider("提案文を生成するスコア閾値", 10, 20, 15)

    st.subheader("AI設定")
    provider = st.selectbox(
        "AIプロバイダー",
        ["gemini", "claude", "openai"],
        format_func=lambda x: {
            "gemini": "Gemini（無料枠あり）",
            "claude": "Claude",
            "openai": "OpenAI GPT-4o mini",
        }[x],
    )
    api_key = st.text_input(
        f"{provider.upper()} APIキー",
        type="password",
        placeholder="APIキーを入力",
    )

    st.subheader("あなたのプロフィール")
    user_profile = st.text_area(
        "自己紹介・スキル・経験など（提案文に反映されます）",
        placeholder="例: Webデザイン歴5年。LP・バナー制作を中心に、WordPressのカスタマイズも対応可能。レスポンシブデザインが得意です。",
        height=140,
    )

# ======================================================================
# メインエリア
# ======================================================================

st.title("クラウドワークス 案件獲得自動化キット")

tab1, tab2, tab3 = st.tabs(["リサーチ", "スコア結果", "提案文"])

# ======================================================================
# Tab 1: リサーチ
# ======================================================================

with tab1:
    st.header("案件リサーチ＆スコアリング")

    col_info, col_btn = st.columns([3, 1])
    with col_info:
        st.markdown(
            f"キーワード: **{' / '.join(keywords)}**  "
            f"&nbsp;&nbsp;閾値: **{threshold}点以上** で提案文生成"
        )
    with col_btn:
        run_btn = st.button("リサーチ開始", type="primary", use_container_width=True)

    if run_btn:
        errors = []
        if not cw_email:
            errors.append("メールアドレスを入力してください")
        if not cw_password:
            errors.append("パスワードを入力してください")
        if not keywords:
            errors.append("キーワードを1つ以上入力してください")

        if errors:
            for e in errors:
                st.error(e)
        else:
            status = st.empty()
            log_box = st.empty()
            logs = []
            _ctx = get_script_run_ctx()

            def update(msg: str):
                add_script_run_ctx(threading.current_thread(), _ctx)
                logs.append(msg)
                log_box.text("\n".join(logs[-8:]))

            with st.spinner("処理中..."):
                try:
                    jobs = run_scraping(
                        email=cw_email,
                        password=cw_password,
                        keywords=keywords,
                        max_pages=max_pages,
                        progress_cb=update,
                    )
                    update(f"{len(jobs)} 件を取得しました。スコアリング中...")

                    config = {"scoring": {"threshold": threshold}}
                    scored = score_all_jobs(jobs, config)
                    st.session_state.scored_results = scored
                    st.session_state.proposals_done = False

                    passed = sum(1 for r in scored if r["scoring"]["passed"])
                    status.success(
                        f"完了。{len(jobs)} 件取得 / "
                        f"スコア {threshold} 点以上: {passed} 件が提案文生成対象です。"
                    )
                except Exception as e:
                    st.error(f"エラー: {e}")
                    st.exception(e)

    if st.session_state.scored_results and not run_btn:
        passed = sum(1 for r in st.session_state.scored_results
                     if r["scoring"]["passed"])
        st.info(
            f"前回の結果: {len(st.session_state.scored_results)} 件 / "
            f"提案文生成対象: {passed} 件  |  "
            f"「スコア結果」タブで確認できます。"
        )

# ======================================================================
# Tab 2: スコア結果
# ======================================================================

with tab2:
    st.header("スコア結果一覧")

    results = st.session_state.scored_results

    if not results:
        st.info("「リサーチ」タブで案件を取得してください。")
    else:
        # フィルター
        col_f, col_s = st.columns(2)
        with col_f:
            view = st.radio(
                "表示",
                ["全件", "合格のみ", "除外のみ"],
                horizontal=True,
            )
        with col_s:
            sort_asc = st.checkbox("スコアを昇順で表示")

        filtered = results
        if view == "合格のみ":
            filtered = [r for r in results if r["scoring"]["passed"]]
        elif view == "除外のみ":
            filtered = [r for r in results if not r["scoring"]["passed"]]
        if sort_asc:
            filtered = sorted(filtered, key=lambda x: x["scoring"]["total"])

        st.caption(f"{len(filtered)} 件表示中")

        # テーブル表示
        table_rows = []
        for r in filtered:
            job = r["job"]
            sc = r["scoring"]
            table_rows.append({
                "合否": "合格" if sc["passed"] else "除外",
                "合計": f"{sc['total']}/20",
                "タイトル": job.get("title", ""),
                "URL": job.get("url", ""),
                "報酬": job.get("budget_text", ""),
                "継続": "継続" if job.get("is_ongoing") else "単発",
                "評価": job.get("client_rating", "-"),
                "時給": f"{sc['scores']['時給']}/4",
                "継続性": f"{sc['scores']['継続性']}/4",
                "CL評価": f"{sc['scores']['クライアント評価']}/4",
                "スキル": f"{sc['scores']['スキル習得度']}/4",
                "精神": f"{sc['scores']['精神コスト']}/4",
            })

        df = pd.DataFrame(table_rows)
        st.dataframe(
            df,
            use_container_width=True,
            column_config={
                "URL": st.column_config.LinkColumn("URL", display_text="開く"),
            },
            hide_index=True,
        )

        # 詳細アコーディオン
        st.divider()
        st.subheader("詳細")

        for r in filtered:
            job = r["job"]
            sc = r["scoring"]
            label = (
                f"{'【合格】' if sc['passed'] else '【除外】'}"
                f" {sc['total']}/20点  {job.get('title', '不明')[:55]}"
            )
            with st.expander(label):
                col_a, col_b = st.columns([2, 1])
                with col_a:
                    url = job.get("url", "")
                    st.markdown(f"**案件URL:** [{url}]({url})")
                    st.markdown(f"**報酬:** {job.get('budget_text', '不明')}")
                    st.markdown(f"**継続:** {'継続案件' if job.get('is_ongoing') else '単発'}")
                    st.markdown(f"**クライアント評価:** {job.get('client_rating', '不明')}")
                    st.markdown(f"**カテゴリ:** {job.get('category', '不明')}")
                with col_b:
                    st.markdown("**スコア内訳**")
                    for item in ["時給", "継続性", "クライアント評価", "スキル習得度", "精神コスト"]:
                        s = sc["scores"].get(item, 0)
                        d = sc["details"].get(item, "")
                        bar = "#" * s + "-" * (4 - s)
                        st.markdown(f"`[{bar}]` **{item}** {s}/4 — {d}")
                    st.markdown(f"**合計: {sc['total']}/20点**")

                if job.get("description"):
                    desc = job["description"]
                    with st.expander("案件説明"):
                        st.text(desc[:600] + ("..." if len(desc) > 600 else ""))

        # Excel ダウンロード
        st.divider()
        if st.button("Excelでダウンロード（提案文なし）", key="dl_score_btn"):
            buf = _to_excel(results)
            st.download_button(
                "ダウンロード",
                data=buf,
                file_name="crowdworks_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_score_file",
            )

# ======================================================================
# Tab 3: 提案文
# ======================================================================

with tab3:
    st.header("提案文生成・確認")

    results = st.session_state.scored_results
    if not results:
        st.info("「リサーチ」タブで案件を取得してください。")
    else:
        config = {"scoring": {"threshold": threshold}}
        passed_results = [r for r in results if r["scoring"]["passed"]]

        if not passed_results:
            st.warning(
                f"スコア {threshold} 点以上の案件が 0 件でした。"
                f"閾値を下げるか、キーワードを変更してください。"
            )
        else:
            st.markdown(
                f"対象案件: **{len(passed_results)} 件**  "
                f"/ AI: **{provider.upper()}**  "
                f"/ APIキー: {'設定済み' if api_key else '未設定'}"
            )

            if not api_key:
                st.warning("サイドバーで APIキーを設定してください。")
            else:
                gen_btn = st.button(
                    f"全 {len(passed_results)} 件の提案文を一括生成",
                    type="primary",
                )

                if gen_btn:
                    log = st.empty()
                    logs = []

                    def gen_progress(msg):
                        logs.append(msg)
                        log.text("\n".join(logs[-5:]))

                    with st.spinner("提案文を生成中..."):
                        generate_all_proposals(
                            st.session_state.scored_results,
                            provider=provider,
                            api_key=api_key,
                            progress_cb=gen_progress,
                            profile=user_profile,
                        )
                        st.session_state.proposals_done = True
                        st.success(f"{len(passed_results)} 件の提案文を生成しました。")
                        st.rerun()

            st.divider()
            for i, r in enumerate(passed_results):
                job = r["job"]
                p = r.get("proposal", {})
                url = job.get("url", "")
                has_proposal = bool(p and (p.get("proposal") or p.get("error")))

                label = f"{r['scoring']['total']}/20点  {job.get('title', '')[:55]}"
                with st.expander(label, expanded=(i == 0 and has_proposal)):
                    st.markdown(f"[案件を開く]({url})  |  スコア: {r['scoring']['total']}/20点")

                    if not has_proposal:
                        # 未生成 → 個別生成ボタン
                        if api_key:
                            if st.button("この案件の提案文を生成", key=f"gen1_{i}"):
                                with st.spinner("生成中..."):
                                    r["proposal"] = generate_proposal(job, provider, api_key, user_profile)
                                    st.session_state.proposals_done = True
                                st.rerun()
                        else:
                            st.info("サイドバーでAPIキーを設定してください。")
                    elif p.get("error"):
                        st.error(f"生成エラー: {p['error']}")
                        if api_key and st.button("再試行", key=f"retry_{i}"):
                            with st.spinner("再生成中..."):
                                r["proposal"] = generate_proposal(job, provider, api_key, user_profile)
                            st.rerun()
                    else:
                        st.caption(f"テンプレート: {p.get('template_name', '')}")
                        edited = st.text_area(
                            "提案文（ここで直接編集できます）",
                            value=p.get("proposal", ""),
                            height=240,
                            key=f"prop_{i}",
                        )
                        if edited != p.get("proposal", ""):
                            r["proposal"]["proposal"] = edited

                        if api_key and st.button("再生成", key=f"regen_{i}"):
                            with st.spinner("再生成中..."):
                                r["proposal"] = generate_proposal(job, provider, api_key, user_profile)
                            st.rerun()

            # ループの外：Excel ダウンロード
            st.divider()
            if st.button("Excelでダウンロード（提案文込み）", type="primary", key="dl_prop_btn"):
                buf = _to_excel(st.session_state.scored_results, include_proposals=True)
                st.download_button(
                    "ダウンロード",
                    data=buf,
                    file_name="crowdworks_results_with_proposals.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_prop_file",
                )


# ======================================================================
# Excel 出力ユーティリティ
# ======================================================================

def _to_excel(scored_results: list, include_proposals: bool = False) -> bytes:
    rows = []
    for r in scored_results:
        job = r["job"]
        sc = r["scoring"]
        p = r.get("proposal", {})
        row = {
            "合否": "合格" if sc["passed"] else "除外",
            "合計スコア": f"{sc['total']}/20",
            "タイトル": job.get("title", ""),
            "URL": job.get("url", ""),
            "報酬": job.get("budget_text", ""),
            "継続案件": "継続" if job.get("is_ongoing") else "単発",
            "CL評価": job.get("client_rating", ""),
            "時給スコア": f"{sc['scores']['時給']}/4",
            "継続性スコア": f"{sc['scores']['継続性']}/4",
            "CL評価スコア": f"{sc['scores']['クライアント評価']}/4",
            "スキルスコア": f"{sc['scores']['スキル習得度']}/4",
            "精神コストスコア": f"{sc['scores']['精神コスト']}/4",
            "検索キーワード": job.get("search_keyword", ""),
        }
        if include_proposals:
            row["テンプレート"] = p.get("template_name", "") if p else ""
            row["提案文"] = p.get("proposal", "") if p else ""
        rows.append(row)

    df = pd.DataFrame(rows)
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="案件一覧")
        ws = writer.sheets["案件一覧"]

        from openpyxl.styles import PatternFill, Font, Alignment
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        bold = Font(bold=True)
        wrap = Alignment(wrap_text=True, vertical="top")

        for cell in ws[1]:
            cell.font = bold

        for row_idx, r in enumerate(scored_results, start=2):
            fill = green if r["scoring"]["passed"] else red
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.alignment = wrap

        # 列幅
        col_widths = {"A": 8, "B": 10, "C": 40, "D": 50, "E": 15,
                      "F": 10, "G": 10, "H": 10, "I": 10, "J": 10,
                      "K": 10, "L": 12, "M": 15, "N": 20, "O": 80}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

    buf.seek(0)
    return buf.read()
