"""
exporter.py - 結果をExcelに出力するモジュール
"""

import os
from datetime import datetime
import pandas as pd


def export_to_excel(scored_results: list, save_folder: str = "./results") -> str:
    """
    スコアリング＆提案文生成結果をExcelに出力

    Returns:
        保存したファイルパス
    """
    os.makedirs(save_folder, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(save_folder, f"crowdworks_results_{timestamp}.xlsx")

    rows = []
    for r in scored_results:
        job = r["job"]
        scoring = r["scoring"]
        proposal_info = r.get("proposal", {})

        row = {
            "タイトル": job.get("title", ""),
            "URL": job.get("url", ""),
            "報酬": job.get("budget_text", ""),
            "継続案件": "◎" if job.get("is_ongoing") else "－",
            "クライアント評価": job.get("client_rating", "不明"),
            "カテゴリ": job.get("category", ""),
            "合計スコア": f"{scoring['total']}/20",
            "🕐 時給": f"{scoring['scores'].get('時給', 0)}/4",
            "🔄 継続性": f"{scoring['scores'].get('継続性', 0)}/4",
            "⭐ クライアント評価": f"{scoring['scores'].get('クライアント評価', 0)}/4",
            "📚 スキル習得": f"{scoring['scores'].get('スキル習得度', 0)}/4",
            "😤 精神コスト": f"{scoring['scores'].get('精神コスト', 0)}/4",
            "合否": "✅ 合格" if scoring["passed"] else "❌ 除外",
            "使用テンプレート": proposal_info.get("template_name", "") if proposal_info else "",
            "提案文": proposal_info.get("proposal", "") if proposal_info else "",
            "検索キーワード": job.get("search_keyword", ""),
        }
        rows.append(row)

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="案件一覧")

        workbook = writer.book
        worksheet = writer.sheets["案件一覧"]

        # 列幅調整
        column_widths = {
            "A": 40,  # タイトル
            "B": 50,  # URL
            "C": 20,  # 報酬
            "D": 10,  # 継続
            "E": 15,  # 評価
            "F": 20,  # カテゴリ
            "G": 12,  # 合計
            "H": 10, "I": 10, "J": 15, "K": 12, "L": 12,  # 各スコア
            "M": 10,  # 合否
            "N": 20,  # テンプレート
            "O": 80,  # 提案文
            "P": 20,  # キーワード
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # 合格案件を緑、除外を赤でハイライト
        from openpyxl.styles import PatternFill, Font
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row_idx, result in enumerate(scored_results, start=2):
            fill = green_fill if result["scoring"]["passed"] else red_fill
            for col_idx in range(1, len(df.columns) + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = fill

        # ヘッダーを太字
        bold_font = Font(bold=True)
        for cell in worksheet[1]:
            cell.font = bold_font

        # 提案文セルの折り返し
        from openpyxl.styles import Alignment
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    return filepath
